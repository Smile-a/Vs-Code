import mysql.connector
from mysql.connector import Error
import redis
import re
import json
import os 

from openpyxl import load_workbook
from tqdm import tqdm
from time import sleep,time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

excelPath = r'C:\Users\wsd\Desktop\MissingData.xlsx'
dataMbUrl = 'https://szzt-sjzt.hbtobacco.cn/dmm/dam-imm-web/model/dataModel'

#用来存储成功修改了的数据模型
redis_client = redis.Redis(host='localhost', port=6379, db=8, password='')

def connect_fetch():
    """ 连接MySQL数据库，获取以cloudmes-开头的数据库中的所有表，存储表名到Redis，
    对于每个唯一的表，获取表结构信息并打印，字段名称转大写 """
    try:
        # 连接数据库
        connection = mysql.connector.connect(
            
        )

        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("成功连接到MySQL数据库，版本：", db_Info)
            
            cursor = connection.cursor()
            cursor.execute("SHOW DATABASES")  # 显示所有数据库
            databases = cursor.fetchall()
            print("所有数据库：")
            for db in databases:
                print(db)
            
            # 过滤以cloudmes-开头的数据库
            cloudmes_dbs = [db[0] for db in databases if db[0].startswith('cloudmes-')]
            print("\n以cloudmes-开头的数据库：")
            for db in cloudmes_dbs:
                print(db)
                
                # 跳过 cloudmes-log 数据库
                if db == 'cloudmes-log':
                    continue
                
                # 连接Redis
                # redis_client = redis.Redis(host='localhost', port=6379, db=8, password='')
                
                # 选择数据库
                cursor.execute(f"USE `{db}`")
                
                # 获取所有表
                cursor.execute("SHOW TABLES")
                tables = cursor.fetchall()
                print(f"\n{db} 中的所有表：")
                for table in tables:
                    # 获取表名全小写
                    tableName = table[0].lower()
                    print(tableName)
                    
                    # 获取表的描述
                    cursor.execute(f"SELECT TABLE_COMMENT FROM information_schema.TABLES WHERE TABLE_SCHEMA='{db}' AND TABLE_NAME='{tableName}'")
                    table_desc = cursor.fetchone()
                    print(f"表描述：{table_desc[0]}")
                
                    # 正则表达式匹配以 '_' 开头后跟任意数量的数字
                    if re.search(r'_\d+$', tableName):
                        # 存在另一个list里，跳过
                        redis_client.sadd("numberList", tableName)
                        continue
                    
                    # 判断表名是否已经存在，不存在则添加
                    if not redis_client.sismember("all_tables", tableName):
                        # 获取表结构  但是没有备注之类
                        # cursor.execute(f"DESCRIBE `{tableName}`")
                        
                        # 用SHOW FULL COLUMNS FROM 就有详细的
                        cursor.execute(f"SHOW FULL COLUMNS FROM `{tableName}`")
                        columns = cursor.fetchall()
                        print(f"\n{tableName} 表结构（字段名称大写）：")
                        for column in columns:
                            print(tuple([field.upper() if isinstance(field, str) else field for field in column]))
                        
                        # 构建一个正则表达式，匹配 '-ha', '-sx', 或 '-gs'
                        pattern = r'(-ha|-sx|-gs)'
                        # 使用 re.sub() 替换掉所有匹配的子串
                        db_cleaned = re.sub(pattern, '', db)
                        # table_desc[0] 为空 就设置为null
                        table_desc = table_desc[0] if table_desc[0] else 'null'
                        # 把表名大写@描述#db名称 组合成为key  db名称去除末尾的 (-ha,-sx,-gs)等 字符串
                        table_desc_key = tableName.upper() + '@' + table_desc + '#' + db_cleaned
                        print(table_desc_key)
                        # 然后把字段转成json格式
                        json_columns = [tuple([field.upper() if isinstance(field, str) else field for field in column]) for column in columns]
                        jsonStr = json.dumps(json_columns)
                        
                        # 添加表名到Redis set集合中
                        redis_client.sadd("all_tables", tableName)
                        # 然后统一放到redis 中 叫做 tableDataModel
                        redis_client.hset("tableDataModel", table_desc_key, jsonStr)
                        
                        # test
                        redis_client.sadd("tablesOnDb", tableName+"@"+db_cleaned)
    except Error as e:
        print("连接或查询失败：", e)
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
            print("MySQL连接已关闭")

def getExcelData() :
    try:
        # 连接Redis
        # redis_client = redis.Redis(host='localhost', port=6379, db=8, password='')
        # 尝试打开excel文件
        workbook = load_workbook(excelPath)
        # 获取活动工作表
        sheet = workbook.active
        # 获取第二列的所有数据
        col_data = [cell.value for cell in sheet['B']]
        # 跳过第一行标题
        col_data = col_data[1:]
        # 成功数量
        success_count = 0
        # 遍历列数据并打印
        for data in col_data:
            if data is not None:  # 检查数据是否不为空                
                codeName = data.upper() + '@'
                # 获取redis中的tableDataModel里面全部数据,模糊匹配
                all_data = redis_client.hgetall("tableDataModel")
                for key, value in all_data.items():
                    if codeName in key.decode():
                        print(f"{key.decode()} 匹配到 {codeName}")
                        # 拆分key 通过 '#'分割得到后面的db名称
                        tableInof, db_name = key.decode().split('#')
                        # tableInof通过@得到数据库名称和描述
                        tableName, tableDesc = tableInof.split('@')
                        # 解析json字符串
                        json_data = json.loads(value.decode())
                        # 遍历json数据
                        for row in json_data:
                            print(row)
                            # 如果第九个字段是'' 就把 codeName+row[0] 拼接成新的key，然后添加到redis中
                            if row[8] == '':
                                new_key = codeName + row[0]
                                redis_client.sadd("tableFieldZdsmIsNullList", new_key)
                        # 匹配成功，计数加一
                        success_count+=1
        print(f"成功匹配 {success_count} 条数据")        
    except FileNotFoundError:
        print(f"文件未找到，请检查路径：{excelPath}")
    except Exception as e:
        print(f"打开文件时发生错误：{e}")
    
def export_data():
    # redis_client = redis.Redis(host='localhost', port=6379, db=8, password='')
    #读取 redis 中 tablesOnDb   Set类型
    all_data = redis_client.smembers("tablesOnDb")
    # 将字节对象转换为字符串
    decoded_data = [item.decode() for item in all_data]
    with open(r'C:\MyProject\Vs-Code\dataModelAuto\data.json', 'w') as f:
        json.dump(decoded_data, f, ensure_ascii=False, indent=4)

def alertMesBox(wait):
    try:
        # 检测一下，当前页面有没有弹窗 一个class="el-message-box__btns" 的消息提醒框，如果有，就获取里面的button元素确定
        messBoxBtn = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'el-message-box__btns')))
        #判断是否存在 并且文本不为空
        if messBoxBtn and messBoxBtn.text != "":
            print("检测到弹窗")
            # 获取里面的button元素确定
            messBoxBtn.find_element(By.TAG_NAME, "button").click()
        else:
            #print("没有检测到弹窗")
            pass
    except Exception as e:
        print("没有找到弹窗，不予理会~")

#开始执行自动化功能
def automationBegins():
    pyFilePath = os.getcwd()
    print("【云MES数据中台-数据模型维护脚本】" + "\t程序路径:" + pyFilePath)
    
    #谷歌浏览器驱动配置
    print("开始操作浏览器~")
    options = webdriver.ChromeOptions()
    # 配置调试端口
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    #加载谷歌浏览器驱动
    chrome = webdriver.Chrome(options=options)
    # 打开指定url
    chrome.get(dataMbUrl)
    print(chrome.title)
    # 使用显示等待确保元素已经加载完成
    wait = WebDriverWait(chrome, 2)
    
    print("成功进入数据模型维护页面")
    # 等会
    sleep(5)
    
    # 定位搜索框
    input_search = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//input[@placeholder='输入关键字进行过滤']")))
    input_search[0].clear()
    input_search[0].send_keys("云MES系统")
    # 自动就查询到了云mes，但是需要选中下才行 找到class="el-card__body"的所有元素
    cloudMesTds = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='el-card__body']//span[@class='el-tooltip item']")))
    # 太多了 
    for tree in cloudMesTds:
        #模糊匹配 是 云MES系统的那一个元素
        if "云MES系统" in tree.text:
            print("找到云mes系统，开始点击~")
            tree.click()
            break

    # 打开excel文件
    workbook = load_workbook(excelPath)
    # 获取活动工作表
    sheet = workbook.active
    # 获取第二列的所有数据
    col_data = [cell.value for cell in sheet['B']]
    # 跳过第一行标题
    col_data = col_data[1:]
    # 成功数量
    success_count = 0
    # 遍历列数据并打印
    for data in col_data:
        if data is not None:  # 检查数据是否不为空
            # 先判断这个data在redis的successList里面，如果有就跳过
            if redis_client.sismember('successList', data.upper()):
                print("模型"+data+"已经存在Redis的successList中，跳过该项")
                continue
            # 拼接码名称 方便匹配redis里的                
            codeName = data.upper() + '@'
            # 判断codeName是否在redis的modelNameIsNullList里，有就跳过后面自己手动处理，或者有了中文名称再说
            if redis_client.sismember('modelNameIsNullList', codeName):
                print("模型"+data+"的code在redis的modelNameIsNullList中，跳过该项")
                continue
            # 获取redis中的tableDataModel里面全部数据,模糊匹配
            all_data = redis_client.hgetall("tableDataModel")
            for key, value in all_data.items():
                if codeName in key.decode():
                    print(f"{key.decode()} 匹配到 {codeName}")
                    # 拆分key 通过 '#'分割得到后面的db名称
                    tableInof, db_name = key.decode().split('#')
                    # tableInof通过@得到数据库名称和描述
                    tableName, tableDesc = tableInof.split('@')
                    
                    # 模型中文名称
                    modelChineseName = tableDesc
                    # 如果中文名称是空的，跳过
                    if modelChineseName == "null":
                        #当有问题的模型，跳过不出来，但是需要把模板的code存到er里面，方便后面手动处理
                        redis_client.sadd('modelNameIsNullList', codeName)
                        continue    
                    # 模型英文名称
                    mxCode = tableName
                    
                    # 检测一下，当前页面有没有弹窗
                    alertMesBox(wait)
                    input_fields = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//input[@placeholder='请输入内容']")))
                    # 忘记了 循环的时候，之前的值没有干掉奥，手动清空
                    input_fields[0].clear()
                    input_fields[1].clear()
                    alertMesBox(wait) #弹窗
                    # 分两次查询 有时候code可以匹配，有的时候 中文可以匹配
                    input_fields[0].send_keys(mxCode)
                    #后面code查不到的时候在用名称查一遍
                    #input_fields[1].send_keys(modelChineseName)
                    
                    # 直接点击查询按钮
                    col_elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".el-col.el-col-24")))
                    # 获取第四个 .el-col el-col-24 元素   就是重置 查询那个div
                    fourth_col_element = col_elements[3]
                    # 找到该元素中的所有 button 元素
                    button_elements = fourth_col_element.find_elements(By.CSS_SELECTOR, "button")
                    # 提前设置一下页面重置按钮，说不定后面会用到
                    backBtnClean = None
                    query_button = None
                    # 遍历所有 button 元素，找到 span 为“查询”的按钮并点击
                    for button in button_elements:
                        span_elements = button.find_elements(By.CSS_SELECTOR, "span")
                        if len(span_elements) > 0 and span_elements[0].text == "重置":
                            # 设置重置按钮
                            backBtnClean = button
                        if len(span_elements) > 0 and span_elements[0].text == "查询":
                            query_button = button
                            button.click()
                            print("点击了'查询'按钮")
                            break
                    sleep(3)
                    # 不出意外一般是成功了,直接定位页面 el-table-body 元素 他们好像是把一个table切成了四个
                    table_body = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".el-table__body")))
                    # 最后一个就是编辑和删除的按钮框
                    table_element = table_body[3]
                    #之后获取这个表格元素里的 tbody
                    tbody_element = table_element.find_element(By.TAG_NAME, "tbody")
                    trRows = tbody_element.find_elements(By.TAG_NAME, "tr")
                    print("表格元素定位成功,共有%d行数据" % len(trRows))
                    # 判断列表是否为空  先只用code匹配一轮，后面再用名称看看，光用名称好像有点问题.
                    if len(trRows) == 0:
                        print("没有找到模型："+mxCode)
                        #execl表中有模型信息，但是通过code在浏览器查询没有数据，保存到webDataNullList下后面手动处理
                        redis_client.sadd('webDataNullList', mxCode)
                        continue
                    #在这里等待一下吧，有数据，但是页面没有加载出来导致的？
                    sleep(2)
                    # 按理就只有一条数据，因为就一个数据模板，所以默认取第一个吧
                    tr_element = None
                    
                    #还是获取结果进行匹配吧，哪怕用code和name一起，它搜索结果还是多个
                    headTable = table_body[2]
                    # 打印headTable所有元素
                    for tr in headTable.find_elements(By.TAG_NAME, "tr"):
                        tds = tr.find_elements(By.TAG_NAME, "td")
                        sjmxybmL = tds[3]
                        if sjmxybmL.text == mxCode:
                            deleteMxCode = mxCode
                            print("excel中的code和页面查询到的数据code完全匹配!!! "+sjmxybmL.text)
                            # 记录一个下标
                            tr_element_num = tds[1].text
                            # 就是要编辑这一行 这样就能定位到code搜索结果返回多个数据里最符合的那行数据了
                            tr_element = trRows[int(tr_element_num) - 1] #trRows里面是下标 num存的是序号，所以要减一
                            break
                    # 这里有一种情况，有模糊匹配的结果，但是T_P开头，匹配不到 那不还是NONe
                    if tr_element == None:
                        print(f"******没有完全匹配CODE的模板，跳过该项数据,模板code:{mxCode},模型名称:{modelChineseName}******")
                        # 这种情况是通过code可以模糊匹配到数据，但是对应的数据表名有可能是T_P开头这种,以及code模糊相似名称不同的，就需要手动处理了
                        redis_client.sadd('TpList', sheet.name)
                        continue
                    
                    # 然后通过class=el-table_1_column_10 is-center  is-hidden el-table__cell 定位到 编辑按钮触发
                    # 查看tr_element所有子元素  10个td
                    #print(tr_element.find_elements(By.TAG_NAME, "td").__len__())
                    # 获取最后的td
                    td_element = tr_element.find_elements(By.TAG_NAME, "td")[9]
                    # 查看td里面的所有子元素
                    #print(td_element.find_elements(By.TAG_NAME, "a").__len__())
                    # 按理就俩a标签，第一个是编辑，第二个是删除，获取第一个编辑按钮，然后触发
                    edit_button = td_element.find_elements(By.TAG_NAME, "a")[0]
                    alertMesBox(wait) # 检测弹出提示框
                    edit_button.click()
                    #太快了没加载出来就会报错
                    sleep(2)
                    
                    # 新增bug，有时候会弹窗说编辑操作提示，要点击确定按钮 class="el-message-box"
                    try:  
                        #或者直接找class=el-button el-button--default el-button--small el-button--primary 的button按钮 有就直接点击
                        alertBoxs = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='el-message-box__wrapper']")))
                        #找到alertBoxs[0]里面的所有button元素
                        alert_buttons = alertBoxs[0].find_elements(By.TAG_NAME, "button")
                        for button in alert_buttons:
                            if button.text == "确定":
                                alert_button = button
                                alert_button.click()
                                print("点击了编辑操作提示弹窗的确定按钮")
                                sleep(3) # 页面要加载一下
                                break
                    except:
                        print("没有编辑操作提示弹窗，不用管")
                    
                    #这时候就要弹出模态框了，进行编辑，把excel的对应数据填写到表单中即可。
                    # 获取父级中的 section 元素
                    section_element = wait.until(EC.presence_of_element_located((By.TAG_NAME, 'section')))
                    # 从section中获取各个对应的输入框即可。
                    # 先获取 class="dam_imm-info_model_edit"的div 这是大编辑form
                    edit_div = section_element.find_element(By.CLASS_NAME, "dam_imm-info_model_edit")
                    # 获取和mt-8同级的div class="drawerBtn" 这是暂存按钮的div容器，里面有个button才是真的暂存按钮
                    drawerBtn_div = edit_div.find_element(By.CLASS_NAME, "drawerBtn")
                    # 获取它下面的 button 元素 这是暂存按钮
                    zcSave_button = drawerBtn_div.find_element(By.TAG_NAME, "button")
                    # 然后获取它下面的 class="mt-8" div 这是主从表的容器
                    main_div = edit_div.find_element(By.CLASS_NAME, "mt-8")
                    # 直接获取main_div里面所有包括子元素里面的 class="el-input__inner"
                    input_elements = main_div.find_elements(By.CLASS_NAME, "el-input__inner")
                    
                    # 进入模型后循环添加字段
                    # 解析json字符串
                    json_data = json.loads(value.decode())
                    # 遍历json数据
                    # for row in json_data:
                    #     print(row)
                    
                    # 遍历所有的行  目前已知 每页6列 行数未知
                    # 字段名	字段类型(长度)	能否为NULL	是否主键    是否默认	''
                    for row in tqdm(json_data):
                        # 暂停1秒
                        sleep(1)
                        # 字段名 统一转为大写
                        zdm = row[0]
                        zdm = zdm.upper()
                        # 字段说明- 也就是字段中文名
                        zdsm = row[8]
                        # 字段类型
                        zdlx = row[1]
                        # 字段长度=row[1]字符()里面的数字   'VARCHAR(72)'   'DECIMAL(19,2)'
                        zdlx_length = re.findall(r'\d+', zdlx)
                        if len(zdlx_length) == 1:
                            zdcd = int(zdlx_length[0])
                            xsws = 0
                        elif len(zdlx_length) == 2:
                            zdcd = int(zdlx_length[0])
                            xsws = int(zdlx_length[1])
                        else:
                            zdcd = 0
                            xsws = 0
                        # 能否为NULL   YES NO
                        nfwn = row[3]
                        # 是否主键  '' PRI
                        sfzj = row[4]
                        # 是否默认
                        sffr = row[5]
                        
                        # 这时候已经获取到了当前行里的所有元素，然后依次进行填充
                        # 先获取当前模态窗里面 id="pane-column" 的元素 这里面会有手动添加按钮和数据列表
                        tabpanel = edit_div.find_element(By.ID, "pane-column")
                        
                        # 获取tabpanel里面的button 一共有五个，获取第四个"手动添加"按钮
                        sleep(1)
                        try:
                            add_button = tabpanel.find_elements(By.TAG_NAME, "button")[3]
                            add_button.click()
                        except Exception as e:
                            print("重新调用手动添加")
                            # 检测提醒    
                            alertMesBox(wait)
                            sleep(1)
                            # 一般都是因为上一个手动编辑的页面框关闭失败导致的，再关闭一次
                            sdxgzdCloseBtn = drawerAddForm.find_element(By.TAG_NAME, "i")
                            sleep(1)
                            #sdxgzdCloseBtn.click()
                            # 使用 JavaScript 执行点击
                            chrome.execute_script("arguments[0].click();", sdxgzdCloseBtn)
                            #然后再次调用 手动新增 按钮
                            add_button = tabpanel.find_elements(By.TAG_NAME, "button")[3]
                            sleep(2)
                            #这里也会弹窗接口超时
                            # 检测提醒    
                            alertMesBox(wait)
                            # 一切正常 点击手动新增按钮
                            add_button.click()
                        
                        #定位弹窗的class="el-drawer__wrapper" 有很多个
                        sleep(2)
                        drawers = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'el-drawer__wrapper')))
                        # 获取最后一个元素 也就是 手动添加的弹窗
                        drawerAddForm = drawers[-1]
                        
                        # 开始输入值 获取所有input标签
                        input_elements = drawerAddForm.find_elements(By.TAG_NAME, "input")
                        # 一共是有11个
                        print("一共有%d个input元素" % len(input_elements))
                        
                        # 检测提醒    
                        alertMesBox(wait)
                        
                        # 开始给对应的输入框赋值 第一个输入框是 源字段名
                        input_elements[0].send_keys(zdm)
                        # 第二个输入框是 字段中文名
                        if zdsm == '':
                            if "DOMAINCOL" == zdm:
                                zdsm = "域"
                            elif "SRL" == zdm:
                                zdsm = "顺序号"
                            elif "STDTEMPLATEID":
                                zdsm = "工艺标准模板树主表ID"
                            elif "PRECISION_STR":
                                zdsm = "精度"
                            else:
                                zdsm = zdm
                                # 记录没有匹配到的字段说明，存到redis中[tableFieldNullList] 表名@字段名
                                redis_client.sadd("tableFieldZdsmIsNullList", mxCode + "@" + zdm)
                        input_elements[1].send_keys(zdsm)
                        # 字段类型截取( 默认是字符串  如果没有(那就是INT DATE DECIMAL这种类型
                        zdlxs = zdlx.split("(")
                        if len(zdlxs) == 2:
                            zdlx = zdlxs[0]
                        else:
                            # 可能就不是varchar(64) 这种类型 也许是DAte或者其他
                            zdlx = zdlxs
                        # 重新拼接zdlx 带中文组装
                        if zdlx == "DECIMAL":
                            zdlx = "高精小数(Decimal)"
                        elif zdlx == "DATETIME":
                            zdlx = "时间(Time)"
                        elif zdlx == "INT" or zdlx == "TINYINT":
                            zdlx = "整数(Integer)"
                        elif zdlx == "VARCHAR":
                            zdlx = "变长字符(Varchar)"
                        # 不太对，字段类型 input是一个下拉框，直接输入值好像不能成功，要如何操作?
                        input_elements[3].send_keys(zdlx)
                        # 直接这样是不行的，鼠标移动就丢失值，需要 键盘下然后加上回车 就可以了
                        input_elements[3].send_keys(Keys.DOWN, Keys.ENTER)
                        # 第五个输入框是 字段长度
                        if zdcd != 'null' and zdcd != "" and zdcd != None and zdcd != 0:
                            try:
                                input_elements[4].send_keys(zdcd)
                            except Exception as e:
                                print("可能是其他类型，但是没有输入值，跳过")
                        # 第六个输入框是 小数位 如果zdlx是decimal 才需要设置
                        if zdlx == "高精小数(Decimal)":
                            # 字段长度一般只有在字段类型是varchar的时候才有用  数字类型一般就是null，那么就改获取数字长度
                            input_elements[4].send_keys(zdcd)
                            input_elements[5].send_keys(xsws)
                        elif zdlx == "整数(Integer)":
                            input_elements[4].send_keys(zdcd)
                        # 能否为NULL 默认是不限制非空的 
                        if nfwn == "NO":
                            # 数据中为NO 表示不允许为空，单击非空约束控件即可 启用
                            fkysInput = input_elements[8]
                            sleep(1)
                            print(fkysInput.is_enabled())
                            # 使用 JavaScript 执行点击
                            chrome.execute_script("arguments[0].click();", fkysInput)
                        # 判断是否是主键 sfzj = PRI 就是主键，其余不管
                        if sfzj == "PRI":
                            # 点击主键按钮
                            zjBtn = input_elements[2]
                            sleep(1)
                            print(zjBtn.is_enabled())
                            chrome.execute_script("arguments[0].click();", zjBtn)
                        #获取里面 class="drawerBtn" 元素里面的 button  这是保存按钮
                        drawerBtnDiv = drawerAddForm.find_element(By.CLASS_NAME, "drawerBtn")
                        saveBtn = drawerBtnDiv.find_element(By.TAG_NAME, "button")
                        saveBtn.click()
                        
                        # 检测页面是否有弹出提示 role="alert" 的元素
                        try:
                            alert_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[role='alert']")))
                            # 提示框里面的p标签 class=el-message__content 的text
                            alert_P = alert_element.find_element(By.CLASS_NAME, "el-message__content")
                            alert_text = alert_P.text
                            #判断不为空的话 就打印出来
                            if alert_text != "":
                                print(alert_text)
                                #print("源字段名已存在，跳过该字段!!!")
                                sdxgzdCloseBtn = drawerAddForm.find_element(By.TAG_NAME, "i")
                                chrome.execute_script("arguments[0].click();", sdxgzdCloseBtn)
                                continue
                        except Exception as e:
                            print("没有捕获到顶部弹窗~")
                            # 判断当前关闭按钮还是不是存在 get_attribute 是否none 状态
                            sdxgzdCloseBtn = drawerAddForm.find_element(By.TAG_NAME, "i")
                            if sdxgzdCloseBtn.aria_role != 'none':
                                chrome.execute_script("arguments[0].click();", sdxgzdCloseBtn)
                    
                    # 数据模型的字段都修改好了 可以直接点击暂存按钮了
                    try:
                        zcSave_button.click()
                    except Exception as e:
                        print("获取调用暂存按钮失败,重新调用")
                        # 一般都是因为上一个手动编辑的页面框关闭失败导致的，再关闭一次
                        sdxgzdCloseBtn = drawerAddForm.find_element(By.TAG_NAME, "i")
                        sleep(1)
                        sdxgzdCloseBtn.click()
                        #然后再次调用 暂存 按钮
                        zcSave_button.click()
                    print("数据模型修改完毕，暂存成功！！！")
                    #之后就可以关闭弹窗了，定位 aria-label="close 数据模型管理" 的button 按钮 点击执行关闭数据模型修改模态框
                    close_button = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "button[aria-label='close 数据模型管理']")))
                    sleep(2)
                    close_button.click()
                    
                    # 这一个sheet页的数据就都处理完毕了，把sheet页的code存储到redis的successList里面，下次循环就可以跳过这个sheet页了
                    redis_client.sadd("successList", deleteMxCode)
                    # 匹配成功，计数加一
                    success_count+=1
                    # 打印分割线
                    print("==================================================================================================================")
    print(f"成功匹配 {success_count} 条数据")    



if __name__ == '__main__':
    # 调用函数 通过mysql获取所有的表和结构到redis中
    # connect_fetch()
    
    # 读取excel里的数据
    # getExcelData()
    
    # 导出redis里tablesOnDb的数据为json文件
    # export_data()
    
    # 调用自动测试函数
    automationBegins()
    
    pass